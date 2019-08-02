from openpyxl import Workbook
from openpyxl import load_workbook
import pandas as pd
from py2neo import *

wb = load_workbook('test.xlsx')
ws = wb.active
rows = []
for row in ws.iter_rows():
    rows.append(row)

graph = Graph("http://localhost:7474", username="yourname", password="yourpassworld")

#建事件名称节点
event_name="S"+str(rows[1][0].value)
s = "match (a:航空安全事件) create (b:"+event_name+"{name:'"+str(rows[1][0].value)+"'}) create (a)-[:包含]->(b)"
graph.run(s)

index = "match (b:"+event_name+")"

#时间节点建立
check_time = "T"+ str(rows[1][1])
find_time = graph.find_one(
    label = check_time
)
if find_time==None:
    construct_time = index + " match (c:时间) create (d:T" + str(rows[1][1].value) +"{name:'"+str(rows[1][1].value)+"'}) create (c)-[:包含]->(d) create (b)-[:属性{name:'时间'}]->(d)"
    graph.run(construct_time)
else:
    construct_time = index + " match (d:T"+str(rows[1][1].value)+") create (b)-[:属性{name:'时间'}]->(d)"
    graph.run(construct_time)

#客机型号节点建立
check_plane = str(rows[1][2].value)
find_plane = graph.find_one(
    label= check_plane
)
if find_plane==None:
    construct_plane = index + " match (g:客机型号) create(h:" + check_plane + "{name:'"+check_plane+"'}) create (g)-[:包含]->(h) create (b)-[:属性{name:'客机型号'}]->(h)"
    graph.run(construct_plane)
else :
    construct_plane = index + "match (h:"+check_plane+") create (b)-[:属性{name:'客机型号'}]->(h)"

#航空公司节点建立
check_airline = str(rows[1][3].value)
find_airline = graph.find_one(
    label= check_airline
)
if find_airline==None:
    construct_airline = index + " match (e:航空公司) create(f:"+ check_airline + "{name:'" + check_airline +"'}) create (e)-[:包含]->(f) create (b)-[:属性{name:'航空公司'}]->(f)"
    graph.run(construct_airline)
else:
    construct_airline = index + "match (f:" + check_airline +") create (b)-[:属性{name:'航空公司'}]->(f)"
    graph.run(construct_airline)

# 航班号节点建立
construct_flight = index + "match (hbh:航班号) create(jhbh:h" +str(rows[1][4].value)+"{name:'"+ str(rows[1][4].value) +"'}) create (hbh)-[:包含]->(jhbh) create (b)-[:属性{name:'航班号'}]->(jhbh)"
graph.run(construct_flight)

#起飞地点节点建立
check_beginPlace = str(rows[1][5].value)
find_beginPlace = graph.find_one(
    label= check_beginPlace
)
if find_beginPlace == None:
    construct_beginPlace = index + "match (qf:起飞地点) create(jqf:"+ check_beginPlace +"{name:'"+ check_beginPlace +"'}) create (qf)-[:包含]->(jqf) create (b)-[:属性{name:'起飞地点'}]->(jqf)"
    graph.run(construct_beginPlace)
else:
    construct_beginPlace = index + "match (qf:起飞地点) match(jqf:"+ check_beginPlace +") create (qf)-[:包含]->(jqf) create (b)-[:属性{name:'起飞地点'}]->(jqf)"
    graph.run(construct_beginPlace)

#降落地点节点建立
check_landPlace = str(rows[1][6].value)
find_landPlace = graph.find_one(
    label=check_landPlace
)
if find_landPlace == None:
    construct_landPlace = index + " match (qf:降落地点) create(jqf:" + check_landPlace + "{name:'"+ check_landPlace +"'}) create (qf)-[:包含]->(jqf) create (b)-[:属性{name:'降落地点'}]->(jqf)"
    graph.run(construct_landPlace)
else:
    construct_landPlace = index + " match (qf:降落地点) match(jqf:" + check_landPlace +") create (qf)-[:包含]->(jqf) create (b)-[:属性{name:'降落地点'}]->(jqf)"
    graph.run(construct_landPlace)

#出事地点节点建立
check_incidentPlace = str(rows[1][7].value)
find_incidentPlace = graph.find_one(
    label=check_incidentPlace
)
if check_incidentPlace == None:
    construct_incidentPlace = index + " match(cs:出事地点) create(jcs:"+ check_incidentPlace + "{name:'" +check_incidentPlace +"'}) create (cs)-[:包含]->(jcs) create (b)-[:属性{name:'出事地点'}]->(jcs)"
    graph.run(construct_incidentPlace)
else:
    construct_incidentPlace = index + " match(cs:出事地点) match (jcs:" + check_incidentPlace +") create (cs)-[:包含]->(jcs) create (b)-[:属性{name:'出事地点'}]->(jcs)"
    graph.run(construct_incidentPlace)

#事件类型节点建立
check_eventType = str(rows[1][8].value)
find_eventType = graph.find_one(
    label = check_eventType
)
if find_eventType ==None:
    construct_eventType = index + "match (sl:事件类型) create (jsl:"+ check_eventType + "{name:'"+ check_eventType +"'}) create (sl)-[:包含]->(jsl) create (b)-[:属性{name:'事件类型'}]->(jsl)"
    graph.run(construct_eventType)
else:
    construct_eventType = index + "match (jsl:"+ check_eventType +") create (b)-[:属性{name:'事件类型'}]->(jsl)"
    graph.run(construct_eventType)

#航线类型节点建立
check_airlineType = str(rows[1][9].value)
if check_airlineType == "国际":
    construct_airlineType = index + "match(jhl:国际) create (b)-[:属性{name:'航线类型'}]->(jhl)"
    graph.run(construct_airlineType)
else:
    construct_airlineType = index + "match(jhl:国内) create (b)-[:属性{name:'航线类型'}]->(jhl)"

#航班类型节点建立
check_flightType = str(rows[1][10].value)
if check_flightType == "客运":
    construct_flightType = index + "match (jhb:客运) create (b)-[:属性{name:'航班类型'}]->(jhb)"
    graph.run(construct_flightType)
else:
    construct_flightType = index + "match (jhb:货运) create (b)-[:属性{name:'航班类型'}]->(jhb)"

#天气情况节点建立
check_weather = str(rows[1][11].value)
find_weater = graph.find_one(
    label=check_weather
)
if find_weater==None:
    construct_weather = index + "match (tq:天气情况) create (jtq:"+ check_weather +"{name:'"+ check_weather +"'}) create (tq)-[:包含]->(jtq) create (b)-[:属性{name:'天气情况'}]->(jtq)"
    graph.run(construct_weather)
else:
    construct_weather = index + "match(jtq:"+ check_weather +") create (b)-[:属性{name:'天气情况'}]->(jtq)"
    graph.run(construct_weather)

#操作阶段节点建立
check_stage = str(rows[1][12].value)
find_stage = graph.find_one(
    label=check_stage
)
if find_stage == None:
    construct_stage = index + "match (cj:操作阶段) create (jcj:" + check_stage +"{name:'"+ check_stage +"'}) create (cj)-[:包含]->(jcj) create (b)-[:属性{name:'操作阶段'}]->(jcj)"
    graph.run(construct_stage)
else:
    construct_stage = index + "match(jcj:"+check_stage +") create (b)-[:属性{name:'操作阶段'}]->(jcj)"
    graph.run(construct_stage)

#原因节点建立
reason = str(rows[1][13].value)
construct_reason = index + "match (rea:原因) create (jrea:"+ reason +"{name:'"+ reason +"'}) create (rea)-[:包含]->(jrea) create (b)-[:属性{name:'原因'}]->(jrea)"
graph.run(construct_reason)

#结果节点建立
result = str(rows[1][15].value)
construct_result = index + "match (resul:结果) create (jresul:"+ result +"{name:'"+ result +"'}) create (resul)-[:包含]->(jresul) create (b)-[:属性{name:'结果'}]->(jresul)"
graph.run(construct_result)

#人员伤亡节点建立
injured = str(rows[1][14].value)
if injured == "无":
    construct_injured = ""
else:
    construct_injured = index + "match (ry:人员伤亡) create (jresul:"+ injured +"{name:'"+ injured +"'}) create (ry)-[:包含]->(jresul) create (b)-[:属性{name:'人员伤亡'}]->(jresul)"
    graph.run(construct_injured)

#事件等级节点建立
grade = str(rows[1][16].value)
find_grade = graph.find_one(
    label=grade
)
if find_grade == None:
    construct_grade = index + " match (dj:等级) create (jdj:"+ grade + "{name:'" + grade +"'}) create (dj)-[:包含]->(jdj) create (b)-[:属性{name:'等级'}]->(jdj)"
    graph.run(construct_grade)
else:
    construct_grade = index + " match (jdj:"+ grade +") create (b)-[:属性{name:'等级'}]->(jdj)"
    graph.run(construct_grade)

print("end!")